from telegram.ext import Updater, MessageHandler, Filters
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import signal
import sys

# ANSI color codes
YELLOW = '\033[93m'
BLUE = '\033[94m'
MAGENTA = '\033[95m'
RESET = '\033[0m'

# Replace with your actual bot token
BOT_TOKEN = 'Telegram bot API key (Get from BotFather in Telegram)'
EXCEL_FILE = 'group_info.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Groups"
        ws.append(['Timestamp', 'Group Name', 'Chat ID'])  # Header row
        wb.save(EXCEL_FILE)

def is_chat_logged(chat_id):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if str(row[2]) == str(chat_id):  # Compare as strings
            return True
    return False

def log_group_info(update, context):
    chat = update.effective_chat

    if chat.type in ['group', 'supergroup']:
        group_name = chat.title
        chat_id = chat.id
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if not is_chat_logged(chat_id):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([timestamp, group_name, str(chat_id)])
            wb.save(EXCEL_FILE)
            print(f"{YELLOW}Logged: [{timestamp}] Group Name: {group_name} | Chat ID: {chat_id}{RESET}")
        else:
            print(f"{BLUE}Already logged: {group_name} ({chat_id}){RESET}")

def print_poll_timestamp(context):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"{MAGENTA}Polling at: {timestamp}{RESET}")

def signal_handler(sig, frame):
    print('\nBot is shutting down gracefully. Bye!')
    sys.exit(0)

def main():
    signal.signal(signal.SIGINT, signal_handler)

    init_excel()
    updater = Updater(BOT_TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(MessageHandler(Filters.all, log_group_info))

    # Add repeating job to print timestamp every 5 seconds
    job_queue = updater.job_queue
    job_queue.run_repeating(print_poll_timestamp, interval=5, first=0)

    # Start polling with poll_interval of 0.5 seconds (adjust as you want)
    updater.start_polling(poll_interval=0.5)
    print("Bot started. Press Ctrl+C to stop.")
    updater.idle()

if __name__ == '__main__':
    main()
